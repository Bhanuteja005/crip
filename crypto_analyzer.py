import requests
import pandas as pd
import time
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import ColorScaleRule
import logging
from pathlib import Path
import json
def fetch_top_cryptocurrencies(limit=50):
    """
    Fetch top 50 cryptocurrencies from CoinGecko API
    """
    url = f"https://api.coingecko.com/api/v3/coins/markets"
    params = {
        "vs_currency": "usd",
        "order": "market_cap_desc",
        "per_page": limit,
        "page": 1,
        "sparkline": "false"
    }
    
    try:
        response = requests.get(url, params=params)
        response.raise_for_status()
        data = response.json()
        
        # Extract required fields
        crypto_data = [{
            "Name": coin["name"],
            "Symbol": coin["symbol"].upper(),
            "Current Price (USD)": coin["current_price"],
            "Market Capitalization": coin["market_cap"],
            "24h Trading Volume": coin["total_volume"],
            "24h Price Change (%)": coin["price_change_percentage_24h"]
        } for coin in data]
        
        return pd.DataFrame(crypto_data)
    
    except requests.RequestException as e:
        print(f"Error fetching data: {e}")
        return None

def analyze_cryptocurrency_data(df):
    if df is None or df.empty:
        return None
    analysis = {
        "Top 5 Cryptocurrencies by Market Cap": df.head(5)[["Name", "Symbol", "Market Capitalization"]].to_string(index=False),
        "Average Price": f"${df['Current Price (USD)'].mean():.2f}",
        "Highest 24h Price Change": df.loc[df['24h Price Change (%)'].idxmax()][["Name", "Symbol", "24h Price Change (%)"]].to_dict(),
        "Lowest 24h Price Change": df.loc[df['24h Price Change (%)'].idxmin()][["Name", "Symbol", "24h Price Change (%)"]].to_dict()
    }
    
    return analysis

def update_excel_sheet(df, analysis=None):
    wb = openpyxl.Workbook()
    
    ws1 = wb.active
    ws1.title = "Cryptocurrency Data"
    
    ws1.cell(row=1, column=1, value=f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center')
    
    price_change_col = df.columns.get_loc("24h Price Change (%)") + 1
    ws1.conditional_formatting.add(
        f'G3:G{len(df)+2}', 
        ColorScaleRule(start_type='min', start_color='FF0000',
                       end_type='max', end_color='00FF00')
    )
    
    if analysis:
        ws2 = wb.create_sheet(title="Analysis")
        for idx, (key, value) in enumerate(analysis.items(), 1):
            ws2.cell(row=idx, column=1, value=key).font = Font(bold=True)
            ws2.cell(row=idx, column=2, value=str(value))
    
    wb.save("cryptocurrency_data.xlsx")
    print("Excel sheet updated successfully!")

def main():
    while True:
        df = fetch_top_cryptocurrencies()
        
        if df is not None:
            analysis = analyze_cryptocurrency_data(df)
            
            update_excel_sheet(df, analysis)
        time.sleep(300)

if __name__ == "__main__":
    main()